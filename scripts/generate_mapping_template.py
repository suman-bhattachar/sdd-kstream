"""Generate templates/mapping.template.xlsx — the analyst-facing feed mapping workbook.

The Field Mappings TARGET zone fully determines the target topic's Avro value schema:
type vocabulary maps 1:1 to Avro (incl. decimal(p,s), enum, logical date/time types),
dotted paths express nested records/arrays, Required drives null unions, Field
Description becomes the Avro field doc, and the Feed sheet's Schema settings block
carries record name / namespace / compatibility.

Dev-side maintainer tool: needs `openpyxl` (pip install openpyxl). Never runs at
install time or inside the agent loop — the stdlib-only rule applies to runtime
scripts, not this generator. Re-run it to regenerate the template after edits.

Usage:  python scripts/generate_mapping_template.py [output.xlsx]
"""

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = sys.argv[1] if len(sys.argv) > 1 else str(
    Path(__file__).resolve().parent.parent / "templates" / "mapping.template.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SOURCE_FILL = PatternFill("solid", fgColor="DDEBF7")   # light blue  — source side
RULE_FILL   = PatternFill("solid", fgColor="FFF2CC")   # light amber — rule
TARGET_FILL = PatternFill("solid", fgColor="E2EFDA")   # light green — target side
SECTION_FILL = PatternFill("solid", fgColor="D9D9D9")
EXAMPLE_FONT = Font(italic=True, color="808080")
LABEL_FONT = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

RULE_TYPES = "CONSTANT,DIRECT,TRANSFORM,CONDITIONAL,LOOKUP,SYSTEM"
TARGET_TYPES = "text,whole number,big whole number,decimal number,yes/no,date,time,timestamp,choice list"
ON_ERROR = "FAIL_TO_DLQ,USE_DEFAULT,LEAVE_EMPTY,SKIP_RECORD"
YES_NO = "Y,N"
COMPAT_MODES = "BACKWARD,FULL,FORWARD,NONE"

wb = Workbook()

# ------------------------------------------------------------------ Instructions
ws = wb.active
ws.title = "Instructions"
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 120

rows = [
    ("title", "Feed Mapping Workbook — How to fill this in"),
    ("", ""),
    ("text", "One workbook = one feed (one source topic mapped to one target topic). Fill the sheets in order:"),
    ("text", "   1. Feed — facts about the feed as a whole (topics, key, owner). The Schema settings block at the bottom is ENGINEERING ONLY — leave it alone."),
    ("text", "   2. Field Mappings — one row per TARGET field. Source columns are on the LEFT, the rule in the middle, the target on the RIGHT."),
    ("text", "   3. Lookups — only if any mapping row uses rule type LOOKUP: define each lookup once here."),
    ("", ""),
    ("head", "The golden rule: EVERY mapping row MUST have a worked example."),
    ("text", "Fill 'Sample Source Value(s)' and 'Expected Target Value' with a real, concrete example. The example is the final word:"),
    ("text", "if your English rule and your example disagree, the engineering side will come back and ask — they will never guess."),
    ("text", "The examples also become the automated tests for the feed."),
    ("", ""),
    ("head", "Rule types (the 'Rule Type' dropdown) and how to write each rule"),
    ("text", "CONSTANT     — the target is always a fixed value.            Write:  Always \"EU\""),
    ("text", "DIRECT       — copy the source field as-is.                   Write:  Copy as-is"),
    ("text", "TRANSFORM    — reshape one or more source fields.             Write:  Last 4 characters of cardNumber   /   Uppercase of countryCode   /   Concatenate firstName and lastName with a space"),
    ("text", "CONDITIONAL  — if/else logic.                                 Write:  IF amount is greater than 10000 THEN \"HIGH\" OTHERWISE \"NORMAL\"  (you can chain: ... OTHERWISE IF ... THEN ...)"),
    ("text", "LOOKUP       — fetch a value from a reference collection.     Write:  Look up <LookupName> using <source field(s)>; take <attribute>   — and define <LookupName> on the Lookups sheet."),
    ("text", "SYSTEM       — generated at processing time, no source.       Write:  System timestamp at processing time   /   Unique ID"),
    ("", ""),
    ("head", "Target types (the 'Target Type' dropdown) — pick exactly one per row"),
    ("text", "text             — free text of any length."),
    ("text", "whole number     — an integer up to about 2 billion."),
    ("text", "big whole number — an integer beyond 2 billion (IDs, counts of everything)."),
    ("text", "decimal number   — exact decimal (money!). MUST also fill 'Precision,Scale', e.g. 18,2 = up to 18 digits, 2 after the point."),
    ("text", "yes/no           — true/false."),
    ("text", "date             — a calendar date without time, e.g. 2026-07-03."),
    ("text", "time             — a time of day without date, e.g. 10:15:00."),
    ("text", "timestamp        — date + time down to milliseconds, e.g. 2026-07-03T10:15:00.000Z."),
    ("text", "choice list      — one value out of a fixed set. MUST also fill 'Allowed Values', e.g. HIGH,NORMAL."),
    ("", ""),
    ("head", "Field names and nested fields (the 'Target Field (path)' column)"),
    ("text", "Field names use letters, digits and underscore only, and start with a letter — e.g. customer_id, riskBand. No spaces, no dashes."),
    ("text", "A dot creates a sub-structure:      customer.city         → a 'customer' block containing 'city'."),
    ("text", "[] creates a repeating list:        charges[].amount      → a list of charge entries, each with an 'amount'."),
    ("text", "                                    tags[]                → a simple list of values (the Target Type applies to each entry)."),
    ("text", "Keep all rows of the same block together (all charges[].* rows next to each other)."),
    ("text", "A block or list is treated as optional only when ALL of its fields are optional (Required = N)."),
    ("", ""),
    ("head", "The two 'default' columns — they are different things"),
    ("text", "On-Error Default  — the value used AT RUNTIME when the rule fails or the source is missing and you chose USE_DEFAULT."),
    ("text", "Schema Default    — a technical default written into the data contract for future compatibility. Normally LEAVE IT BLANK;"),
    ("text", "                    optional fields are handled automatically. Only fill it if your engineer asks for a specific value."),
    ("", ""),
    ("head", "What you never touch"),
    ("text", "The 'Schema settings' block on the Feed sheet is engineering-owned: the design phase confirms record name, namespace,"),
    ("text", "timestamp precision and compatibility. Leave every pre-filled value exactly as it is."),
    ("", ""),
    ("head", "When something is missing or fails ('If Missing / On Error' dropdown)"),
    ("text", "FAIL_TO_DLQ  — the whole record is rejected to the dead-letter queue (use for mandatory fields)."),
    ("text", "USE_DEFAULT  — use the value you put in the 'On-Error Default' column."),
    ("text", "LEAVE_EMPTY  — the target field is left empty (only allowed if Required = N)."),
    ("text", "SKIP_RECORD  — the record is silently dropped (rare — use only when agreed with engineering)."),
    ("", ""),
    ("head", "Style rules"),
    ("text", "• Write rules in plain English using the patterns above. One row per target field — never describe two fields in one row."),
    ("text", "• Use exact source field names as they appear in the source data (case-sensitive)."),
    ("text", "• Write n/a in the Source / Sample / On-Error columns when the rule has no source (SYSTEM and CONSTANT rows)."),
    ("text", "• Fill 'Field Description' for every row — one sentence saying what the field means. It goes into the data contract."),
    ("text", "• Don't merge cells. Don't add or reorder columns. Add extra rows freely."),
    ("text", "• Grey italic rows in the sheets are worked examples — replace them with your real content."),
]
r = 1
for kind, text in rows:
    c = ws.cell(row=r, column=2, value=text)
    if kind == "title":
        c.font = Font(bold=True, size=14)
    elif kind == "head":
        c.font = Font(bold=True, size=11)
    r += 1

# ------------------------------------------------------------------ Feed
ws = wb.create_sheet("Feed")
ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 55
ws.column_dimensions["C"].width = 75

ws.cell(row=1, column=1, value="Field").font = HEADER_FONT
ws.cell(row=1, column=2, value="Value").font = HEADER_FONT
ws.cell(row=1, column=3, value="Help").font = HEADER_FONT
for col in (1, 2, 3):
    ws.cell(row=1, column=col).fill = HEADER_FILL

# (label, prefilled value, help) — None label = section divider row
feed_rows = [
    ("Feed name", "", "Short unique name for this feed, e.g. payments-enriched"),
    ("Description / business purpose", "", "One or two sentences: what this feed is for"),
    ("SOURCE — topic name", "", "The Kafka topic this feed reads from"),
    ("SOURCE — schema (subject or link)", "", "Schema Registry subject name, or a link to the source schema"),
    ("SOURCE — one record represents", "", "e.g. 'one payment transaction'"),
    ("TARGET — topic name", "", "The Kafka topic this feed writes to"),
    ("TARGET — key field(s)", "", "Which target field(s) form the message key (plain string key)"),
    ("TARGET — key join separator", "-", "Only if several key fields: the character between them, e.g. C-1001-EU"),
    ("Records can be dropped/filtered?", "", "Y/N — if Y, describe the filter condition in plain English below"),
    ("Filter condition (if any)", "", "e.g. 'Ignore records where status is CANCELLED'"),
    ("Dead-letter queue topic", "", "Leave blank to use the standard convention (engineering fills this)"),
    ("Business owner (analyst)", "", "Name + contact of the person who owns this mapping"),
    ("Version / date", "", "e.g. v1.0 — 2026-07-03"),
    (None, "Schema settings — ENGINEERING ONLY: confirmed by the design phase (/sdd-architect). Analysts: do not edit.", None),
    ("SCHEMA — target record name", "PaymentsEnriched", "Technical name of the data contract. PascalCase, no spaces — usually the feed name"),
    ("SCHEMA — namespace", "com.yourorg.feeds", "Reverse-domain grouping — your engineer knows the org convention"),
    ("SCHEMA — record description", "", "One sentence describing the record as a whole — goes into the data contract"),
    ("SCHEMA — timestamp precision", "timestamp-millis", "Org default. Do not change without engineering"),
    ("SCHEMA — compatibility mode", "BACKWARD", "How future schema changes are checked. Org default is BACKWARD"),
]
r = 2
compat_cell = None
for entry in feed_rows:
    label, val, help_ = entry
    if label is None:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        c = ws.cell(row=r, column=1, value=val)
        c.font = Font(bold=True, size=11)
        c.fill = SECTION_FILL
        r += 1
        continue
    ws.cell(row=r, column=1, value=label).font = LABEL_FONT
    ws.cell(row=r, column=2, value=val)
    h = ws.cell(row=r, column=3, value=help_)
    h.font = EXAMPLE_FONT
    for col in (1, 2, 3):
        ws.cell(row=r, column=col).border = BORDER
        ws.cell(row=r, column=col).alignment = WRAP
    if label.startswith("SCHEMA — compatibility"):
        compat_cell = f"B{r}"
    r += 1

d = DataValidation(type="list", formula1=f'"{COMPAT_MODES}"', allow_blank=True, showDropDown=False)
d.add(compat_cell)
ws.add_data_validation(d)
ws.freeze_panes = "A2"

# ------------------------------------------------------------------ Field Mappings
ws = wb.create_sheet("Field Mappings")

# (header, width, zone)  zone: src | rule | tgt | none
cols = [
    ("Source Field(s)",          20, "src"),
    ("Sample Source Value(s)",   22, "src"),
    ("Rule Type",                14, "rule"),
    ("Rule (plain English)",     44, "rule"),
    ("If Missing / On Error",    18, "rule"),
    ("On-Error Default",         14, "rule"),
    ("Target Field (path)",      22, "tgt"),
    ("Target Type",              15, "tgt"),
    ("Precision,Scale",          13, "tgt"),
    ("Allowed Values",           18, "tgt"),
    ("Required",                 10, "tgt"),
    ("Schema Default",           13, "tgt"),
    ("Field Description",        34, "tgt"),
    ("Expected Target Value",    22, "tgt"),
    ("Notes",                    28, "none"),
]
ZONE_FILL = {"src": SOURCE_FILL, "rule": RULE_FILL, "tgt": TARGET_FILL}

# zone banner (row 1) + headers (row 2)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)    # source
ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=6)    # rule
ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=14)   # target
for c0, text in ((1, "SOURCE (what comes in)"), (3, "RULE (what happens)"),
                 (7, "TARGET (what goes out — defines the data contract)")):
    c = ws.cell(row=1, column=c0, value=text)
    c.font = Font(bold=True, size=11)
    c.alignment = Alignment(horizontal="center")
for idx, (_, _, zone) in enumerate(cols, start=1):
    if zone != "none":
        ws.cell(row=1, column=idx).fill = ZONE_FILL[zone]

for idx, (header, width, zone) in enumerate(cols, start=1):
    c = ws.cell(row=2, column=idx, value=header)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = WRAP
    c.border = BORDER
    ws.column_dimensions[get_column_letter(idx)].width = width

# worked example rows (grey italic — analyst replaces)
# src, sample, rule type, rule, on error, on-error default,
# target path, target type, prec/scale, allowed, req, schema default, description, expected, notes
examples = [
    ("n/a", "n/a", "SYSTEM", "System timestamp at processing time", "n/a", "",
     "processed_at", "timestamp", "", "", "Y", "",
     "Time the record was processed by the pipeline", "2026-07-03T10:15:00.000Z", "Set by the pipeline"),
    ("n/a", "n/a", "CONSTANT", 'Always "EU"', "n/a", "",
     "region_code", "text", "", "", "Y", "",
     "Fixed region marker for this feed", "EU", ""),
    ("custId", "C-1001", "DIRECT", "Copy as-is", "FAIL_TO_DLQ", "",
     "customer_id", "text", "", "", "Y", "",
     "Unique customer identifier from the source system", "C-1001", "Mandatory — reject if absent"),
    ("cardNumber", "4111222233334444", "TRANSFORM", "Last 4 characters of cardNumber", "LEAVE_EMPTY", "",
     "card_last4", "text", "", "", "N", "",
     "Last four digits of the card number", "4444", ""),
    ("amount", "12000.50", "TRANSFORM", "Copy amount rounded to 2 decimal places", "FAIL_TO_DLQ", "",
     "net_amount", "decimal number", "18,2", "", "Y", "",
     "Transaction amount in feed currency", "12000.50", ""),
    ("amount", "12000.50", "CONDITIONAL",
     'IF amount is greater than 10000 THEN "HIGH" OTHERWISE "NORMAL"', "FAIL_TO_DLQ", "",
     "risk_band", "choice list", "", "HIGH,NORMAL", "Y", "",
     "Risk classification of the transaction", "HIGH", ""),
    ("charges[].chargeType", "FEE", "DIRECT", "Copy as-is for each charge", "LEAVE_EMPTY", "",
     "charges[].type", "text", "", "", "N", "",
     "Type of each charge line", "FEE", "One target charge entry per source charge"),
    ("charges[].chargeAmount", "3.50", "DIRECT", "Copy as-is for each charge", "LEAVE_EMPTY", "",
     "charges[].amount", "decimal number", "18,2", "", "N", "",
     "Amount of each charge line", "3.50", ""),
    ("branchId", "B-77", "LOOKUP", "Look up Branches using branchId; take name", "USE_DEFAULT", "UNKNOWN",
     "branch_name", "text", "", "", "N", "",
     "Human-readable branch name", "Munich Central", "Branches defined on Lookups sheet"),
]
for r_i, row in enumerate(examples, start=3):
    for c_i, val in enumerate(row, start=1):
        c = ws.cell(row=r_i, column=c_i, value=val)
        c.font = EXAMPLE_FONT
        c.alignment = WRAP
        c.border = BORDER

# blank formatted rows for the analyst
for r_i in range(len(examples) + 3, 60):
    for c_i in range(1, len(cols) + 1):
        c = ws.cell(row=r_i, column=c_i)
        c.border = BORDER
        c.alignment = WRAP

# dropdowns
def dv(formula, ranges):
    d = DataValidation(type="list", formula1=f'"{formula}"', allow_blank=True, showDropDown=False)
    for rng in ranges:
        d.add(rng)
    ws.add_data_validation(d)

dv(RULE_TYPES,   ["C3:C200"])
dv(ON_ERROR,     ["E3:E200"])
dv(TARGET_TYPES, ["H3:H200"])
dv(YES_NO,       ["K3:K200"])

ws.freeze_panes = "A3"

# ------------------------------------------------------------------ Lookups
ws = wb.create_sheet("Lookups")
lk_cols = [
    ("Lookup Name",                      18),
    ("Collection / reference data set",  28),
    ("Match: source field(s) -> lookup key field(s)", 40),
    ("Attributes used",                  26),
    ("If not found",                     18),
    ("Expected miss rate",               18),
    ("Notes",                            36),
]
for idx, (header, width) in enumerate(lk_cols, start=1):
    c = ws.cell(row=1, column=idx, value=header)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = WRAP
    c.border = BORDER
    ws.column_dimensions[get_column_letter(idx)].width = width

lk_example = ("Branches", "branches (MongoDB)", "branchId -> branch_id", "name, city",
              "USE_DEFAULT (per field)", "< 1%", "Reference data owned by the Branch team")
for c_i, val in enumerate(lk_example, start=1):
    c = ws.cell(row=2, column=c_i, value=val)
    c.font = EXAMPLE_FONT
    c.alignment = WRAP
    c.border = BORDER
for r_i in range(3, 20):
    for c_i in range(1, len(lk_cols) + 1):
        c = ws.cell(row=r_i, column=c_i)
        c.border = BORDER
        c.alignment = WRAP

d = DataValidation(type="list", formula1=f'"{ON_ERROR}"', allow_blank=True, showDropDown=False)
d.add("E2:E100")
ws.add_data_validation(d)
ws.freeze_panes = "A2"

wb.save(OUT)
print(f"written: {OUT}")
