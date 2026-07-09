"""Convert an analyst feed-mapping workbook (.xlsx) into specs/<feed>/mapping.md.

Analyst-phase intake tool, run by /sdd-analyst. Validates ONLY analyst-owned content
(rules, examples, types, lookups); the Feed sheet's Schema settings block passes
through UNVALIDATED — it is design-phase input owned by /sdd-architect (see
knowledge/mapping-rules.md). Emits human-readable tables plus a fenced JSON block
that downstream tooling (scripts/generate_avro_schema.py) consumes.

Needs openpyxl (pip install openpyxl) — runs at feed intake on an engineer's
machine, never at install time (the stdlib-only rule covers install/runtime hooks).

Usage:
  python scripts/convert_mapping.py <workbook.xlsx> <output mapping.md> [--source-schema src.avsc]

Exit codes: 0 = converted, 1 = validation errors (listed per row; nothing written).
"""

import datetime as _dt
import hashlib
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

RULE_TYPES = {"CONSTANT", "DIRECT", "TRANSFORM", "CONDITIONAL", "LOOKUP", "SYSTEM"}
TARGET_TYPES = {"text", "whole number", "big whole number", "decimal number",
                "yes/no", "date", "time", "timestamp", "choice list"}
ON_ERROR = {"FAIL_TO_DLQ", "USE_DEFAULT", "LEAVE_EMPTY", "SKIP_RECORD"}
NO_SOURCE_RULES = {"CONSTANT", "SYSTEM"}
AVRO_NAME = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")
LOOKUP_REF = re.compile(r"look\s*up\s+(\w+)", re.IGNORECASE)

FM_COLS = ["source", "sample", "rule_type", "rule", "on_error", "on_error_default",
           "target", "target_type", "precision_scale", "allowed_values", "required",
           "schema_default", "description", "expected", "notes"]


def cell(ws, r, c):
    v = ws.cell(row=r, column=c).value
    return str(v).strip() if v is not None else ""


def read_feed_sheet(wb):
    ws = wb["Feed"]
    facts, schema = {}, {}
    for r in range(2, ws.max_row + 1):
        label, value = cell(ws, r, 1), cell(ws, r, 2)
        if not label or "ENGINEERING ONLY" in label:
            continue
        if label.startswith("SCHEMA"):
            schema[label] = value
        else:
            facts[label] = value
    return facts, schema


def read_mappings(wb):
    ws = wb["Field Mappings"]
    rows = []
    for r in range(3, ws.max_row + 1):
        values = {k: cell(ws, r, i + 1) for i, k in enumerate(FM_COLS)}
        if not any(values.values()):
            continue
        values["_row"] = r
        rows.append(values)
    return rows


def read_lookups(wb):
    ws = wb["Lookups"]
    lookups = {}
    for r in range(2, ws.max_row + 1):
        name = cell(ws, r, 1)
        if not name:
            continue
        lookups[name] = {
            "name": name, "collection": cell(ws, r, 2), "match": cell(ws, r, 3),
            "attributes": cell(ws, r, 4), "if_not_found": cell(ws, r, 5),
            "expected_miss_rate": cell(ws, r, 6), "notes": cell(ws, r, 7),
        }
    return lookups


def source_schema_fields(path):
    """Top-level + dotted field names of an .avsc, arrays flattened as name[]."""
    schema = json.loads(Path(path).read_text(encoding="utf-8"))
    names = set()

    def walk(fields, prefix):
        for f in fields:
            t = f["type"]
            if isinstance(t, list):  # union — take the non-null branch
                t = next((x for x in t if x != "null"), t[0])
            name = prefix + f["name"]
            names.add(name)
            if isinstance(t, dict) and t.get("type") == "array":
                items = t["items"]
                names.add(name + "[]")
                if isinstance(items, dict) and items.get("type") == "record":
                    walk(items["fields"], name + "[].")
            elif isinstance(t, dict) and t.get("type") == "record":
                walk(t["fields"], name + ".")

    walk(schema.get("fields", []), "")
    return names


def validate(facts, rows, lookups, src_fields):
    errs = []

    def err(row, msg):
        errs.append(f"row {row['_row']} ({row.get('target') or '?'}): {msg}")

    for label in ["Feed name", "SOURCE — topic name", "TARGET — topic name",
                  "TARGET — key field(s)"]:
        if not facts.get(label):
            errs.append(f"Feed sheet: '{label}' is empty")

    seen_targets = set()
    for row in rows:
        rt, tt = row["rule_type"], row["target_type"]
        if rt not in RULE_TYPES:
            err(row, f"Rule Type '{rt}' not one of {sorted(RULE_TYPES)}")
            continue
        if not row["target"]:
            err(row, "Target Field (path) is empty")
            continue
        if row["target"] in seen_targets:
            err(row, "duplicate Target Field (path)")
        seen_targets.add(row["target"])
        for seg in row["target"].split("."):
            if not AVRO_NAME.match(seg.replace("[]", "")):
                err(row, f"path segment '{seg}' is not a legal field name")
        if tt not in TARGET_TYPES:
            err(row, f"Target Type '{tt}' not one of {sorted(TARGET_TYPES)}")
        if row["required"] not in ("Y", "N"):
            err(row, "Required must be Y or N")
        if not row["description"]:
            err(row, "Field Description is mandatory")
        if not row["expected"]:
            err(row, "Expected Target Value (the worked example) is mandatory")
        if not row["rule"]:
            err(row, "Rule (plain English) is empty")
        # source / sample / on-error, by rule class
        if rt in NO_SOURCE_RULES:
            pass  # n/a or blank both fine
        else:
            if not row["source"] or row["source"].lower() == "n/a":
                err(row, f"{rt} rule needs Source Field(s)")
            if not row["sample"] or row["sample"].lower() == "n/a":
                err(row, f"{rt} rule needs Sample Source Value(s)")
            if row["on_error"] not in ON_ERROR:
                err(row, f"If Missing / On Error '{row['on_error']}' not one of {sorted(ON_ERROR)}")
            elif row["on_error"] == "USE_DEFAULT" and not row["on_error_default"]:
                err(row, "USE_DEFAULT chosen but On-Error Default is empty")
            elif row["on_error"] == "LEAVE_EMPTY" and row["required"] == "Y":
                err(row, "LEAVE_EMPTY is only allowed when Required = N")
        # type-specific columns
        if tt == "decimal number":
            if not re.match(r"^\d+\s*,\s*\d+$", row["precision_scale"]):
                err(row, "decimal number needs Precision,Scale like 18,2")
        elif tt == "choice list":
            symbols = [s.strip() for s in row["allowed_values"].split(",") if s.strip()]
            if not symbols:
                err(row, "choice list needs Allowed Values")
            for s in symbols:
                if not AVRO_NAME.match(s):
                    err(row, f"allowed value '{s}' is not a legal enum symbol")
        # lookup reference
        if rt == "LOOKUP":
            m = LOOKUP_REF.search(row["rule"])
            if not m:
                err(row, "LOOKUP rule must say 'Look up <Name> using ...'")
            elif m.group(1) not in lookups:
                err(row, f"lookup '{m.group(1)}' is not defined on the Lookups sheet")
        # source fields exist in the source schema (only when a schema was given)
        if src_fields and rt not in NO_SOURCE_RULES:
            for f in re.split(r"[,+]", row["source"]):
                f = f.strip()
                if f and f.lower() != "n/a" and f not in src_fields:
                    err(row, f"source field '{f}' not found in the source schema")
    return errs


def emit(out_path, workbook_path, facts, schema, rows, lookups):
    sha = hashlib.sha256(Path(workbook_path).read_bytes()).hexdigest()[:16]
    today = _dt.date.today().isoformat()
    feed = facts.get("Feed name", "unknown")

    def table(headers, data):
        lines = ["| " + " | ".join(headers) + " |",
                 "|" + "|".join("---" for _ in headers) + "|"]
        for d in data:
            lines.append("| " + " | ".join(str(x).replace("|", "\\|") for x in d) + " |")
        return "\n".join(lines)

    fm_table = table(
        ["Source", "Rule Type", "Rule", "Target", "Type", "Req", "Example (in -> out)"],
        [[r["source"], r["rule_type"], r["rule"], r["target"], r["target_type"],
          r["required"], f"{r['sample']} -> {r['expected']}"] for r in rows])
    lk_table = table(
        ["Name", "Collection", "Match", "Attributes", "If not found"],
        [[l["name"], l["collection"], l["match"], l["attributes"], l["if_not_found"]]
         for l in lookups.values()]) if lookups else "_none_"

    canonical = {"feed": facts, "schema_settings": schema,
                 "fields": [{k: v for k, v in r.items() if k != "_row"} for r in rows],
                 "lookups": list(lookups.values())}

    md = f"""---
feed: {feed}
status: draft
source_workbook: {Path(workbook_path).name}
workbook_sha256: {sha}
converted: {today}
---
# Feed Mapping — {feed}

Converted from the analyst workbook; regenerate with `scripts/convert_mapping.py` after any
workbook change — never hand-edit this file (the workbook is the analyst-facing original).

## Feed facts
{table(["Field", "Value"], sorted(facts.items()))}

## Field mappings
{fm_table}

## Lookups
{lk_table}

## Schema settings (design phase — NOT validated at intake)
Owned by /sdd-architect per `knowledge/mapping-rules.md`; the analyst never fills these.
{table(["Setting", "Value"], sorted(schema.items()))}

## Machine-readable mapping
```json
{json.dumps(canonical, indent=2, ensure_ascii=False)}
```
"""
    Path(out_path).write_text(md, encoding="utf-8", newline="\n")


def main(argv):
    if len(argv) < 3:
        print(__doc__)
        return 1
    workbook_path, out_path = argv[1], argv[2]
    src_fields = None
    if "--source-schema" in argv:
        src_fields = source_schema_fields(argv[argv.index("--source-schema") + 1])

    wb = load_workbook(workbook_path)
    for sheet in ("Feed", "Field Mappings", "Lookups"):
        if sheet not in wb.sheetnames:
            print(f"ERROR: workbook has no '{sheet}' sheet — is this the mapping template?")
            return 1
    facts, schema = read_feed_sheet(wb)
    rows = read_mappings(wb)
    lookups = read_lookups(wb)

    errs = validate(facts, rows, lookups, src_fields)
    if errs:
        print(f"INVALID — {len(errs)} error(s); nothing written:")
        for e in errs:
            print("  - " + e)
        return 1

    emit(out_path, workbook_path, facts, schema, rows, lookups)
    print(f"VALID — {len(rows)} field mappings, {len(lookups)} lookup(s) -> {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
